import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from sklearn.metrics import cohen_kappa_score, f1_score
import os
import json
import numpy as np
from OCI import OCI_from_data
from Stats import plot_result_table, matrix_from_predictions

data_git = os.path.dirname(__file__)+'/../'
TT_labels = json.load(open(data_git+'TT_labels.json'))

def compute_agreement(annotator, periphery=False):
    
    if not os.path.isdir(annotator):
        os.mkdir(annotator)

    TT_labels = json.load(open(data_git+'TT_labels.json'))
    items = ['key'] + TT_labels['main_items'] 
    
    df_flo = pd.read_csv(data_git+'/Annotations/table.csv', index_col=0)
    df_flo = df_flo[df_flo['Shalini']=='test'][items]
    df_other = pd.read_csv(f'{data_git}/Annotations/intergrader/table_{annotator}.csv', index_col=0)[items]
    
    diffs, big_diffs = [], []
    
    common_keys = set(df_flo['key']).intersection(set(df_other['key']))
    print(f'{len(common_keys)} / {len(df_flo)}')
    df_flo = df_flo[df_flo['key'].isin(common_keys)].sort_values(by='key').reset_index(drop=True)
    df_other = df_other[df_other['key'].isin(common_keys)].sort_values(by='key').reset_index(drop=True)
    
    kappas, agrees, ocis, agrees_ok, f1s = [], [], [], [], []
    
    
    for j, item in enumerate(items[1:]):
        
        preds, labels = [], []
        preds_scores, labels_scores = [], []
        
        total, egal, ok = 0, 0, 0
        for i in range(len(df_flo['key'])):
            
            a, b = str(df_other[item][i]), str(df_flo[item][i])
            if a != 'not assessable' and b != 'not assessable':
                preds.append(a), labels.append(b)
                a_score, b_score = TT_labels[item].index(a), TT_labels[item].index(b)
                preds_scores.append(a_score), labels_scores.append(b_score)
                ok_score = int(abs(a_score-b_score)<2)
                
                if a != b:
                    df_other.loc[i, item] = a+'----'+b
                    if ok_score:
                        diffs += [(i, j)]
                    else:
                        big_diffs += [(i, j)]
                
                ok += ok_score
                egal += int(a==b)
                total += 1
            
        matrix_from_predictions(labels_scores, preds_scores, item, annotator)    
        ocis.append(1-OCI_from_data(item, preds, labels))
        kappas.append(cohen_kappa_score(labels, preds))
        agrees.append(egal/total)
        agrees_ok.append(ok/total)
        f1s.append(f1_score(labels_scores, preds_scores, average='weighted'))
    
    df_table = pd.DataFrame({'1-OCI': ocis, 'accuracy': agrees, 'f1-weighted': f1s, 'kappa': kappas}, index=items[1:])
    _items = ['optic disk hyperfluorescence', 'macular edema', 'vascular leakage', 'capillary leakage']
    _items = TT_labels['main_items'] if periphery else _items
    plot_result_table(df_table.loc[_items], annotator)
        
    wb = Workbook()
    ws = wb.active
    for j, item in enumerate(items[1:]):
        ws.cell(row=3, column=j+2).value = item
        ws.cell(row=1, column=j+2).value = round(kappas[j], 3)
        ws.cell(row=2, column=j+2).value = round(100*agrees[j], 1)
    ws.cell(row=1, column=1).value = "Cohen's Kappa"
    ws.cell(row=2, column=1).value = "% agreement"
    ws.cell(row=3, column=1).value = "key"
    for row in df_other.itertuples(index=False):
        ws.append(row)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    blue_fill = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    green_fill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    for (i, j) in diffs:
        ws.cell(row=i+4, column=j+2).fill = orange_fill
    for (i, j) in big_diffs:
        ws.cell(row=i+4, column=j+2).fill = red_fill
    for j in range(len(items)):
        ws.cell(row=3, column=j+1).fill = blue_fill
        for i in range(2):
            ws.cell(row=i+1, column=j+1).fill = green_fill
    for i in range(3):
        ws.cell(row=i+1, column=1).fill = blue_fill
    wb.save(f'{data_git}/Metrics/{annotator}/comparison.xlsx')
    
for annotator in ['Teodora', 'Muriel', 'Shalini']:
    print('computing agreement', annotator, '...')
    compute_agreement(annotator, periphery=1)

folder = os.path.dirname(__file__)+'/../../Classification/results/predictions'
for item in TT_labels['main_items']:
    labels = np.load(folder+'/test_'+item+'_labels.npy')
    preds = np.load(folder+'/test_'+item+'_preds.npy')
    matrix_from_predictions(labels, preds, item, 'model')